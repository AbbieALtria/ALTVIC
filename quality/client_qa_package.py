#!/usr/bin/env python3
"""
Client QA Package Generator
File:     src/quality/client_qa_package.py
Version:  1.0.1  —  local JSON tracking, zero ViciDial DB writes
Purpose:  Build daily QA packages for clients — random recordings per agent,
          AI-scored, packaged as ZIP with PDFs, MP3s, transcripts, Excel summary.

Tracking: History stored locally in exports/client_packages/qa_package_history.json
          Nothing is created or modified on the ViciDial database.
"""

import os
import sys
import json
import random
import shutil
import zipfile
from datetime import datetime, date, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo   # stdlib Python 3.9+  (no extra install needed)

# All dates and times are Eastern Standard Time — matches ViciDial call_date column
EST = ZoneInfo("America/New_York")

def now_est() -> datetime:
    """Current datetime in EST/EDT."""
    return datetime.now(EST)

def today_est() -> date:
    """Current date in EST/EDT."""
    return now_est().date()

def yesterday_est() -> date:
    return today_est() - timedelta(days=1)

sys.path.insert(0, str(Path(__file__).parent.parent))

from core.database import db   # read-only ViciDial queries only
from utils.colors import Colors, print_header, print_success, print_error, \
                         print_info, print_warning, print_color

from quality.ai_assistant import (
    get_whisper_model, download_recording, transcribe_audio,
    analyze_transcript, detect_ghost_call, generate_pdf_report,
    get_checkpoints, GHOST_APPLICABLE_CHECKPOINTS,
    get_confidence_text, get_rating_text,
    TEMP_DIR, EXPORTS_DIR,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Local paths — nothing touches ViciDial ────────────────────────────────────
PACKAGES_DIR = Path(__file__).parent.parent / "exports" / "client_packages"
PACKAGES_DIR.mkdir(parents=True, exist_ok=True)
HISTORY_FILE = PACKAGES_DIR / "qa_package_history.json"

DEFAULT_RECORDINGS_PER_AGENT = 3
MIN_CALL_DURATION = 60


# =============================================================================
# Local JSON history  (zero DB writes)
# =============================================================================

def _load_history():
    try:
        if HISTORY_FILE.exists():
            return json.loads(HISTORY_FILE.read_text(encoding='utf-8'))
    except Exception:
        pass
    return []


def _save_history(history):
    HISTORY_FILE.write_text(
        json.dumps(history, indent=2, default=str),
        encoding='utf-8'
    )


def record_package(pkg_meta):
    history = _load_history()
    history.append(pkg_meta)
    _save_history(history)


def show_package_history():
    history = _load_history()
    if not history:
        print_warning("No package history yet.")
        return

    print_header("📋  PACKAGE HISTORY  (local)", Colors.CYAN)
    print(f"  {'#':<4} {'Date':<12} {'Campaigns':<25} {'Agents':<7} {'Calls':<6} {'ZIP'}")
    print("-" * 80)
    for i, h in enumerate(reversed(history[-20:]), 1):
        print(f"  {i:<4} {h.get('request_date','?'):<12} "
              f"{h.get('campaign_ids','?')[:24]:<25} "
              f"{h.get('total_agents',0):<7} "
              f"{h.get('total_calls',0):<6} "
              f"{h.get('zip_filename','?')}")
    print("-" * 80)
    print(f"  Total on record: {len(history)}")


# =============================================================================
# ViciDial read-only queries
# =============================================================================

def get_campaigns():
    results = db.execute_query("""
        SELECT DISTINCT campaign_id FROM vicidial_closer_log
        WHERE call_date >= DATE_SUB(NOW(), INTERVAL 30 DAY)
        ORDER BY campaign_id
    """) or []
    return [r['campaign_id'] for r in results if r['campaign_id']]


def get_agents_for_campaign(campaign_id, target_date=None):
    """
    Return agents who had calls on target_date (or last 30 days if no date given).
    Filtering by target_date avoids 'no calls found' for every agent.
    """
    if target_date:
        results = db.execute_query("""
            SELECT DISTINCT a.user, u.full_name
            FROM vicidial_closer_log c
            LEFT JOIN vicidial_agent_log a ON c.uniqueid = a.uniqueid
            LEFT JOIN vicidial_users u ON a.user = u.user
            WHERE c.campaign_id = %s
              AND DATE(c.call_date) = %s
              AND c.length_in_sec >= 60
              AND a.user IS NOT NULL
              AND a.user NOT IN ('6668','6666','6667','QA','MGR')
              AND a.user NOT LIKE '%%17177027118%%'
            ORDER BY a.user
        """, (campaign_id, target_date)) or []
    else:
        results = db.execute_query("""
            SELECT DISTINCT a.user, u.full_name
            FROM vicidial_agent_log a
            LEFT JOIN vicidial_users u ON a.user = u.user
            WHERE a.campaign_id = %s
              AND a.event_time >= DATE_SUB(NOW(), INTERVAL 30 DAY)
              AND a.user NOT IN ('6668','6666','6667','QA','MGR')
              AND a.user NOT LIKE '%%17177027118%%'
            ORDER BY a.user
        """, (campaign_id,)) or []
    return [{'user': r['user'],
             'name': str(r.get('full_name') or r['user'])}
            for r in results if r.get('user')]


def get_random_calls_for_agent(campaign_id, agent_user, target_date, n=3, min_dur=60):
    results = db.execute_query("""
        SELECT c.uniqueid, c.call_date, c.campaign_id,
               c.phone_number, c.length_in_sec, c.queue_seconds,
               a.user AS agent_user, u.full_name AS agent_name
        FROM vicidial_closer_log c
        LEFT JOIN vicidial_agent_log a ON c.uniqueid = a.uniqueid
        LEFT JOIN vicidial_users u ON a.user = u.user
        WHERE DATE(c.call_date) = %s
          AND c.campaign_id = %s
          AND a.user = %s
          AND c.length_in_sec >= %s
        ORDER BY RAND()
        LIMIT %s
    """, (target_date, campaign_id, agent_user, min_dur, n * 4)) or []

    valid = []
    for r in results:
        rec = db.execute_query("""
            SELECT filename, location FROM recording_log
            WHERE filename LIKE %s AND DATE(start_time) = %s LIMIT 1
        """, (f"%{r['phone_number']}%", target_date))
        if rec and rec[0].get('filename'):
            r['filename'] = rec[0]['filename']
            r['location'] = rec[0].get('location', '')
            valid.append(r)
        if len(valid) >= n * 2:
            break

    random.shuffle(valid)
    return valid[:n * 2]


# =============================================================================
# Helpers
# =============================================================================

def safe_name(s):
    return "".join(c if c.isalnum() or c in (' ', '_', '-') else '_'
                   for c in str(s)).strip().replace(' ', '_')


# =============================================================================
# Per-call processor
# =============================================================================

def process_single_call(call, checkpoints, pkg_agent_dir, date_str):
    phone         = call.get('phone_number', 'unknown')
    duration      = call.get('length_in_sec', 0) or 0
    agent_name    = call.get('agent_name') or call.get('agent_user') or 'Unknown'
    campaign      = call.get('campaign_id', 'Unknown')
    call_date_obj = call.get('call_date', now_est())
    call_date_str = (call_date_obj.strftime('%Y%m%d_%H%M%S')
                     if hasattr(call_date_obj, 'strftime') else date_str)

    print_info(f"    {phone}  dur={duration//60}:{duration%60:02d}")

    audio_path = download_recording(call['uniqueid'], phone, call_date_obj)
    if not audio_path:
        print_warning("    No recording — skip")
        return None

    transcript = transcribe_audio(audio_path)
    if not transcript:
        print_warning("    Transcription failed — skip")
        try:
            audio_path.unlink()
        except Exception:
            pass
        return None

    ghost_info = detect_ghost_call(transcript, duration)
    call_type  = ghost_info['call_type']

    if call_type == 'GHOST':
        print_color(f"    GHOST: {ghost_info['reason'][:55]} — skip", Colors.YELLOW)
        try:
            audio_path.unlink()
        except Exception:
            pass
        return None

    # Score
    analysis = analyze_transcript(transcript)
    analysis['transcript_preview'] = transcript[:500] + ("..." if len(transcript) > 500 else "")
    analysis['call_type']    = call_type
    analysis['ghost_reason'] = ghost_info['reason']

    na_checkpoints = set()
    ai_scores      = {}
    for cp in checkpoints:
        cp_id   = cp['checkpoint_id']
        max_pts = cp['max_points']
        order   = cp.get('display_order', 0)
        if call_type == 'SHORT_REVIEW' and order not in GHOST_APPLICABLE_CHECKPOINTS:
            ai_scores[cp_id] = 0
            na_checkpoints.add(cp_id)
        else:
            raw = analysis['scores'].get(cp_id, 7)
            ai_scores[cp_id] = min(max_pts, max(0, round(raw * max_pts / 10)))

    total_max   = sum(c['max_points'] for c in checkpoints
                      if c['checkpoint_id'] not in na_checkpoints)
    total_score = sum(v for cid, v in ai_scores.items()
                      if cid not in na_checkpoints)
    percentage  = (total_score / total_max * 100) if total_max > 0 else 0

    base = (f"QC_{safe_name(campaign)}_{safe_name(agent_name)}"
            f"_{call_date_str}_{str(phone)[-4:]}")

    # PDF
    pdf_dest = pkg_agent_dir / f"{base}.pdf"
    pdf_path = generate_pdf_report(call, ai_scores, checkpoints,
                                   total_score, total_max, percentage,
                                   analysis, None, na_checkpoints)
    if pdf_path and pdf_path.exists():
        shutil.copy(pdf_path, pdf_dest)

    # MP3
    mp3_dest = pkg_agent_dir / f"{base}.mp3"
    if audio_path.exists():
        shutil.copy(audio_path, mp3_dest)

    # Transcript TXT
    txt_dest = pkg_agent_dir / f"{base}_transcript.txt"
    txt_dest.write_text(transcript, encoding='utf-8')

    try:
        audio_path.unlink()
    except Exception:
        pass

    return {
        'agent_name':  agent_name,
        'agent_user':  call.get('agent_user', ''),
        'campaign':    campaign,
        'call_date':   call_date_obj,
        'phone':       phone,
        'duration':    duration,
        'call_type':   call_type,
        'total_score': total_score,
        'total_max':   total_max,
        'percentage':  percentage,
        'confidence':  analysis.get('confidence', 70),
        'notes':       analysis.get('notes', ''),
        'pdf_file':    pdf_dest.name,
        'mp3_file':    mp3_dest.name,
    }


# =============================================================================
# Excel summary
# =============================================================================

def generate_summary_excel(results, output_path, package_date):
    if not EXCEL_AVAILABLE or not results:
        return False

    wb  = Workbook()
    ws  = wb.active
    ws.title = "QA Summary"

    hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', fgColor='1F4E79')
    hdr_aln   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aln  = Alignment(horizontal='left',   vertical='center')
    ctr_aln   = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='BFBFBF')
    med       = Side(style='medium', color='1F4E79')
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    good_f    = PatternFill('solid', fgColor='E2EFDA')
    warn_f    = PatternFill('solid', fgColor='FFF2CC')
    bad_f     = PatternFill('solid', fgColor='FFE0E0')
    ghost_f   = PatternFill('solid', fgColor='F2F2F2')
    tot_f     = PatternFill('solid', fgColor='D6E4F0')

    ws.merge_cells('A1:L1')
    ws['A1'] = (f"Altria Operations — Client QA Package  |  "
                f"Date: {package_date}  |  "
                f"Generated: {now_est().strftime('%Y-%m-%d %H:%M')} EST")
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    headers    = ['Campaign','Agent','Call Date','Phone','Duration',
                  'Call Type','Score','Max','%','Rating','Confidence%','AI Notes']
    col_widths = [14,22,18,14,10,14,8,6,8,28,13,40]
    for col,(h,w) in enumerate(zip(headers,col_widths),1):
        c = ws.cell(row=2, column=col, value=h)
        c.font=hdr_font; c.fill=hdr_fill; c.alignment=hdr_aln; c.border=bdr
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'A3'

    row = 3
    prev_agent = None
    for r in sorted(results, key=lambda x:(x['agent_name'], str(x['call_date']))):
        pct      = r['percentage']
        is_ghost = r['call_type'] in ('GHOST','SHORT_REVIEW')
        dm, ds   = divmod(r['duration'], 60)
        cl       = get_column_letter

        if prev_agent and r['agent_name'] != prev_agent:
            for col in range(1,13):
                ws.cell(row=row,column=col).border = Border(
                    left=thin,right=thin,bottom=thin,top=med)
        prev_agent = r['agent_name']

        row_data = [
            r['campaign'], r['agent_name'],
            r['call_date'].strftime('%Y-%m-%d %H:%M')
              if hasattr(r['call_date'],'strftime') else str(r['call_date']),
            str(r['phone'])[-10:],
            f"{dm}:{ds:02d}", r['call_type'],
            r['total_score'], r['total_max'],
            f"={cl(7)}{row}/{cl(8)}{row}",
            (get_rating_text(pct).split(' - ')[-1]
             if ' - ' in get_rating_text(pct) else get_rating_text(pct)),
            round(r['confidence'],1), r['notes'],
        ]
        for col,val in enumerate(row_data,1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = bdr
            c.alignment = ctr_aln if col in (1,5,6,7,8,9,11) else left_aln
            if col==9:
                c.number_format='0.0%'
            if col==6 and is_ghost:
                c.font=Font(name='Arial',size=9,italic=True,color='888888')

        rf = ghost_f if is_ghost else good_f if pct>=80 else warn_f if pct>=60 else bad_f
        for col in range(1,13):
            ws.cell(row=row,column=col).fill = rf
        row += 1

    for col in range(1,13):
        ws.cell(row=row,column=col).fill=tot_f
        ws.cell(row=row,column=col).border=bdr
    ws.cell(row=row,column=1,value='TOTALS').font=Font(name='Arial',bold=True,size=10)
    ws.cell(row=row,column=7,value=f'=SUM(G3:G{row-1})').font=Font(name='Arial',bold=True)
    ws.cell(row=row,column=8,value=f'=SUM(H3:H{row-1})').font=Font(name='Arial',bold=True)
    pc=ws.cell(row=row,column=9,value=f'=G{row}/H{row}')
    pc.number_format='0.0%'; pc.font=Font(name='Arial',bold=True)
    kc=ws.cell(row=row,column=11,value=f'=AVERAGE(K3:K{row-1})')
    kc.number_format='0.0'; kc.font=Font(name='Arial',bold=True)

    wb.save(str(output_path))
    return True


# =============================================================================
# ZIP
# =============================================================================

def zip_package(pkg_dir, zip_path):
    with zipfile.ZipFile(zip_path,'w',zipfile.ZIP_DEFLATED) as zf:
        for f in sorted(pkg_dir.rglob('*')):
            if f.is_file():
                zf.write(f, f.relative_to(pkg_dir.parent))
    return zip_path


# =============================================================================
# Interactive UI
# =============================================================================

def select_campaigns_multi():
    campaigns = get_campaigns()
    if not campaigns:
        print_error("No campaigns found in the last 30 days")
        return []
    print("\n📋 SELECT CAMPAIGN(S)  [numbers separated by commas, or A for all]:")
    print("-"*60)
    for i,c in enumerate(campaigns,1):
        print(f"  {i:3}. {c}")
    print("  A. All campaigns")
    print("  0. Cancel")
    choice = input(f"\n{Colors.CYAN}Choice: {Colors.RESET}").strip().lower()
    if choice=='0': return []
    if choice=='a': return campaigns
    selected=[]
    for p in choice.split(','):
        p=p.strip()
        if p.isdigit() and 1<=int(p)<=len(campaigns):
            selected.append(campaigns[int(p)-1])
    if not selected:
        print_error("No valid selection")
    return selected


def get_last_call_date(campaign_id):
    """Return the most recent call date for a campaign — helps QA pick the right date."""
    result = db.execute_query("""
        SELECT DATE(MAX(call_date)) AS last_date
        FROM vicidial_closer_log
        WHERE campaign_id = %s
          AND call_date >= DATE_SUB(NOW(), INTERVAL 7 DAY)
    """, (campaign_id,)) or []
    if result and result[0].get('last_date'):
        return str(result[0]['last_date'])
    return None


def select_package_date(campaigns=None):
    yesterday = yesterday_est()
    today     = today_est()

    # Show last call date per campaign as a hint
    if campaigns:
        print("\n  Last call date per campaign:")
        for c in campaigns:
            last = get_last_call_date(c)
            hint = last if last else "no recent calls"
            print(f"    {c:<20} → {hint}")

    print(f"\n📅 SELECT DATE  (all times EST):")
    print(f"  1. Yesterday ({yesterday})")
    print(f"  2. Today     ({today})")
    print("  3. Custom    (YYYY-MM-DD)")
    choice = input(f"\n{Colors.CYAN}Choice (Enter = yesterday): {Colors.RESET}").strip()
    if choice == '2': return today
    if choice == '3':
        raw = input("  Date (YYYY-MM-DD): ").strip()
        try:
            return datetime.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            print_warning("Invalid date — using yesterday")
    return yesterday


def select_recordings_per_agent():
    raw=input(f"\n{Colors.CYAN}Recordings per agent (default {DEFAULT_RECORDINGS_PER_AGENT}): {Colors.RESET}").strip()
    if raw.isdigit() and 1<=int(raw)<=10: return int(raw)
    return DEFAULT_RECORDINGS_PER_AGENT


# =============================================================================
# Main
# =============================================================================

def show_client_qa_package():
    print_header("📦  CLIENT QA PACKAGE GENERATOR", Colors.CYAN)
    print("\n  Builds daily ZIP package for clients.")
    print("  PDF + MP3 + transcript per call | Excel summary | local history log")
    print("  Nothing written to ViciDial database.\n")
    print("  1. Build new package")
    print("  2. View package history")
    print("  0. Back")

    menu=input(f"\n{Colors.CYAN}Choice: {Colors.RESET}").strip()
    if menu=='0': return
    if menu=='2':
        show_package_history()
        input("\nPress Enter to continue...")
        return
    if menu!='1': return

    campaigns   = select_campaigns_multi()
    if not campaigns:
        input("\nPress Enter to continue...")
        return

    pkg_date    = select_package_date(campaigns)
    n_per_agent = select_recordings_per_agent()

    print(f"\n{'='*60}")
    print(f"  Campaigns : {', '.join(campaigns)}")
    print(f"  Date      : {pkg_date}")
    print(f"  Per agent : {n_per_agent} recording(s)")
    print(f"{'='*60}")
    if input(f"\n{Colors.CYAN}Start? (y/N): {Colors.RESET}").strip().lower()!='y':
        input("\nPress Enter to continue...")
        return

    # ── Pre-flight: verify calls exist on selected date BEFORE loading Whisper ──
    print_info("Checking available calls for selected date...")
    preflight_agents = {}
    for c in campaigns:
        agents_check = get_agents_for_campaign(c, pkg_date)
        preflight_agents[c] = agents_check

    total_available = sum(len(v) for v in preflight_agents.values())

    if total_available == 0:
        print("\n" + "="*60)
        print_error("  ❌ No calls found on this date for any selected campaign.")
        print("="*60)
        for c in campaigns:
            last = get_last_call_date(c)
            if last:
                print_info(f"  {c:<20} last call date → {last}")
        print_warning("\n  Please go back and select the correct date.")
        input("\nPress Enter to continue...")
        return

    # Show what we found
    print_success(f"  Found calls on {pkg_date}:")
    for c, agents in preflight_agents.items():
        if agents:
            print_info(f"    {c}: {len(agents)} agent(s)")
        else:
            last = get_last_call_date(c)
            hint = f" (last calls: {last})" if last else ""
            print_warning(f"    {c}: no calls on this date{hint}")
    # ─────────────────────────────────────────────────────────────────────────

    checkpoints=get_checkpoints()
    if not checkpoints:
        print_error("No checkpoints in database")
        input("\nPress Enter to continue...")
        return

    print_info("Loading AI model...")
    try:
        get_whisper_model()
    except Exception as e:
        print_error(f"Could not load Whisper: {e}")
        input("\nPress Enter to continue...")
        return

    camp_label = "_".join(safe_name(c) for c in campaigns)
    pkg_name   = f"ClientQA_{camp_label}_{pkg_date.strftime('%Y%m%d')}"
    pkg_dir    = PACKAGES_DIR/pkg_name
    if pkg_dir.exists(): shutil.rmtree(pkg_dir)
    pkg_dir.mkdir(parents=True)

    all_results=[]; total_agents=0; total_calls=0; total_ghost=0

    for campaign in campaigns:
        print_header(f"📋  {campaign}", Colors.MAGENTA)
        agents = preflight_agents.get(campaign, [])
        if not agents:
            print_warning(f"  No agents with calls on {pkg_date} for {campaign}")
            continue
        print_info(f"  {len(agents)} agent(s)")

        for agent in agents:
            label=agent['name']
            print(f"\n  👤 {label}")
            candidates=get_random_calls_for_agent(
                campaign,agent['user'],pkg_date,n_per_agent,MIN_CALL_DURATION)
            if not candidates:
                print_warning(f"    No calls on {pkg_date}")
                continue

            agent_dir=pkg_dir/safe_name(label)
            agent_dir.mkdir(exist_ok=True)
            scored=0
            for call in candidates:
                if scored>=n_per_agent: break
                result=process_single_call(call,checkpoints,agent_dir,
                                           pkg_date.strftime('%Y%m%d'))
                if result is None:
                    total_ghost+=1
                    continue
                all_results.append(result)
                scored+=1; total_calls+=1

            if scored==0:
                try: agent_dir.rmdir()
                except Exception: pass
                print_warning("    All candidates ghost/failed — skipped")
            else:
                print_success(f"    {scored}/{n_per_agent} done")
                total_agents+=1

    # Guard: don't create an empty ZIP
    if total_calls == 0:
        print_error("\n  No calls were processed — package not created.")
        print_info("  Possible reasons:")
        print_info("    • Wrong date selected (calls may be on a different date)")
        print_info("    • All candidates were ghost calls")
        print_info("    • Recordings not found in recording_log")
        # Clean up empty package folder
        try:
            shutil.rmtree(pkg_dir)
        except Exception:
            pass
        input("\nPress Enter to continue...")
        return

    # Excel
    xlsx_path=pkg_dir/f"ClientQA_Summary_{camp_label}_{pkg_date.strftime('%Y%m%d')}.xlsx"
    if all_results:
        print_info("\nGenerating Excel summary...")
        if generate_summary_excel(all_results,xlsx_path,pkg_date.strftime('%Y-%m-%d')):
            print_success(f"  {xlsx_path.name}")
        else:
            print_warning("  openpyxl unavailable — Excel skipped")

    # ZIP
    zip_path=PACKAGES_DIR/f"{pkg_name}.zip"
    print_info("Assembling ZIP...")
    zip_package(pkg_dir,zip_path)
    zip_mb=zip_path.stat().st_size/(1024*1024) if zip_path.exists() else 0

    # Save to local JSON history — zero ViciDial writes
    record_package({
        "request_date":  str(pkg_date),
        "campaign_ids":  ", ".join(campaigns),
        "created_at":    now_est().strftime('%Y-%m-%d %H:%M:%S EST'),
        "n_per_agent":   n_per_agent,
        "total_agents":  total_agents,
        "total_calls":   total_calls,
        "ghost_skipped": total_ghost,
        "zip_filename":  zip_path.name,
        "zip_size_mb":   round(zip_mb,2),
        "status":        "COMPLETE" if total_calls>0 else "EMPTY",
        "calls":[{
            "agent":    r['agent_name'],
            "campaign": r['campaign'],
            "phone":    str(r['phone'])[-10:],
            "call_date":str(r['call_date']),
            "call_type":r['call_type'],
            "score":    f"{r['total_score']}/{r['total_max']}",
            "pct":      round(r['percentage'],1),
        } for r in all_results],
    })

    print("\n"+"="*60)
    print_color("  📦 PACKAGE COMPLETE", Colors.GREEN)
    print(f"  Campaigns  : {', '.join(campaigns)}")
    print(f"  Date       : {pkg_date}")
    print(f"  Agents     : {total_agents}")
    print(f"  Calls      : {total_calls}  (ghost/failed skipped: {total_ghost})")
    if zip_path.exists():
        print(f"  ZIP        : {zip_path.name}  ({zip_mb:.1f} MB)")
    print(f"  History    : {HISTORY_FILE.name}  (local)")
    print("="*60)

    if input(f"\n{Colors.CYAN}Open output folder? (y/N): {Colors.RESET}").strip().lower()=='y':
        try: os.startfile(str(PACKAGES_DIR))
        except Exception: print_info(f"Folder: {PACKAGES_DIR}")

    input("\nPress Enter to continue...")


if __name__=="__main__":
    show_client_qa_package()
