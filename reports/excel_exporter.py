#!/usr/bin/env python3
# =============================================================================
# File:         excel_exporter.py
# Version:      2.0.0
# Date:         2026-03-07
# Description:  Complete export system - CSV and Excel exports with formatting
# Location:     D:/Altria_Ops/reports/excel_exporter.py
# =============================================================================

import pandas as pd
from datetime import datetime, timedelta
import sys
from pathlib import Path
import os
import csv

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.database import db
from utils.colors import Colors, print_color, print_header, print_success, print_error, print_warning
from utils.formatter import sec_to_hms

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print_warning("openpyxl not installed. Install with: pip install openpyxl")

# =============================================================================
# Helper Functions
# =============================================================================

def ensure_exports_dir():
    """Ensure exports directory exists"""
    exports_dir = Path(__file__).parent / 'exports'
    exports_dir.mkdir(exist_ok=True)
    return exports_dir

def get_date_range():
    """Get date range from user"""
    print("\nSelect date range:")
    print("  1. Today")
    print("  2. Yesterday")
    print("  3. Last 7 days")
    print("  4. Last 30 days")
    print("  5. Last 90 days")
    print("  6. Custom range")
    
    choice = input("\nChoice (1-6): ").strip()
    
    end_date = datetime.now()
    
    if choice == '1':
        start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        range_text = "Today"
    elif choice == '2':
        start_date = (datetime.now() - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=1)
        range_text = "Yesterday"
    elif choice == '3':
        start_date = datetime.now() - timedelta(days=7)
        range_text = "Last 7 days"
    elif choice == '4':
        start_date = datetime.now() - timedelta(days=30)
        range_text = "Last 30 days"
    elif choice == '5':
        start_date = datetime.now() - timedelta(days=90)
        range_text = "Last 90 days"
    elif choice == '6':
        start_str = input("Start date (YYYY-MM-DD): ").strip()
        end_str = input("End date (YYYY-MM-DD): ").strip()
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_str, '%Y-%m-%d') + timedelta(days=1)
            range_text = f"{start_str} to {end_str}"
        except ValueError:
            print_error("Invalid date format. Using last 7 days.")
            start_date = datetime.now() - timedelta(days=7)
            range_text = "Last 7 days"
    else:
        start_date = datetime.now() - timedelta(days=7)
        range_text = "Last 7 days"
    
    return start_date, end_date, range_text

def get_campaigns():
    """Get list of campaigns to export"""
    try:
        query = """
        SELECT DISTINCT campaign_id
        FROM vicidial_closer_log
        WHERE call_date >= DATE_SUB(NOW(), INTERVAL 30 DAY)
        ORDER BY campaign_id
        """
        campaigns = db.execute_query(query)
        campaign_list = [c['campaign_id'] for c in campaigns] if campaigns else []
        
        if not campaign_list:
            print_warning("No campaigns found with recent activity")
            return None
        
        print("\n📋 Available Campaigns:")
        print("-" * 80)
        
        # Display in columns
        col_width = 25
        cols = 3
        
        for i, camp in enumerate(campaign_list, 1):
            display = f"{i:3}. {camp}"
            print(display.ljust(col_width), end="")
            if i % cols == 0:
                print()
        
        if len(campaign_list) % cols != 0:
            print()
        
        print("-" * 80)
        print("Enter campaign numbers (comma-separated), range (e.g., 1-5), or 'all':")
        choice = input("> ").strip().lower()
        
        if choice == 'all' or choice == '':
            return campaign_list
        else:
            selected = []
            for part in choice.split(','):
                part = part.strip()
                if '-' in part:
                    try:
                        start, end = map(int, part.split('-'))
                        for i in range(start, min(end, len(campaign_list)) + 1):
                            selected.append(campaign_list[i-1])
                    except:
                        pass
                elif part.isdigit():
                    idx = int(part) - 1
                    if 0 <= idx < len(campaign_list):
                        selected.append(campaign_list[idx])
            return selected if selected else campaign_list
            
    except Exception as e:
        print_error(f"Error loading campaigns: {e}")
        return None

def format_excel_file(filepath):
    """Apply professional formatting to Excel file"""
    try:
        wb = openpyxl.load_workbook(filepath)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Alternate row colors
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        wb.save(filepath)
        
    except Exception as e:
        print_warning(f"Could not apply formatting: {e}")

# =============================================================================
# CSV Export Functions
# =============================================================================

def export_calls_to_csv():
    """Export raw call data to CSV"""
    print_header("📊 EXPORT CALLS TO CSV", Colors.CYAN)
    
    # Get date range
    start_date, end_date, range_text = get_date_range()
    
    # Get campaigns
    campaigns = get_campaigns()
    if not campaigns:
        return
    
    print(f"\n📊 Exporting call data for {range_text}...")
    
    # Build query
    placeholders = ','.join(['%s'] * len(campaigns))
    query = f"""
    SELECT 
        c.call_date,
        c.campaign_id,
        c.user,
        u.full_name,
        c.phone_number,
        c.status,
        c.length_in_sec,
        c.queue_seconds,
        c.term_reason,
        c.uniqueid
    FROM vicidial_closer_log c
    LEFT JOIN vicidial_users u ON c.user = u.user
    WHERE c.call_date BETWEEN %s AND %s
      AND c.campaign_id IN ({placeholders})
    ORDER BY c.call_date DESC
    LIMIT 100000
    """
    
    params = [start_date, end_date] + campaigns
    
    try:
        results = db.execute_query(query, params)
        
        if not results:
            print_warning("No data found for selected criteria")
            return
        
        # Create filename
        exports_dir = ensure_exports_dir()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = exports_dir / f"calls_export_{timestamp}.csv"
        
        # Write CSV
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            if results:
                writer = csv.DictWriter(f, fieldnames=results[0].keys())
                writer.writeheader()
                writer.writerows(results)
        
        print_success(f"✅ Exported {len(results)} calls to: {filename}")
        print(f"   File size: {os.path.getsize(filename):,} bytes")
        
        # Ask to open folder
        open_choice = input("\nOpen exports folder? (y/N): ").strip().lower()
        if open_choice == 'y':
            os.startfile(exports_dir)
            
    except Exception as e:
        print_error(f"Export failed: {e}")

def export_agents_to_csv():
    """Export agent performance to CSV"""
    print_header("📊 EXPORT AGENT PERFORMANCE TO CSV", Colors.CYAN)
    
    # Get date range
    start_date, end_date, range_text = get_date_range()
    
    print(f"\n📊 Exporting agent data for {range_text}...")
    
    query = """
    SELECT 
        a.user,
        u.full_name,
        COUNT(DISTINCT a.uniqueid) as total_calls,
        SUM(CASE WHEN a.talk_sec >= 5 THEN 1 ELSE 0 END) as answered,
        SUM(CASE WHEN a.status IN ('SALE','YPSALE','UPSELL','CROSSSELL') THEN 1 ELSE 0 END) as sales,
        SUM(a.talk_sec) as total_talk_sec,
        AVG(CASE WHEN a.talk_sec >= 5 THEN a.talk_sec END) as avg_talk_sec,
        SUM(a.pause_sec) as total_pause_sec,
        COUNT(DISTINCT DATE(a.event_time)) as days_active
    FROM vicidial_agent_log a
    LEFT JOIN vicidial_users u ON a.user = u.user
    WHERE a.event_time BETWEEN %s AND %s
    GROUP BY a.user
    ORDER BY total_calls DESC
    """
    
    try:
        results = db.execute_query(query, (start_date, end_date))
        
        if not results:
            print_warning("No agent data found for selected period")
            return
        
        # Format results for export
        formatted_results = []
        for r in results:
            calls = r['total_calls'] or 0
            answered = r['answered'] or 0
            sales = r['sales'] or 0
            
            formatted_results.append({
                'User': r['user'],
                'Full Name': r['full_name'] or 'Unknown',
                'Total Calls': calls,
                'Answered': answered,
                'Answer Rate %': round(answered / calls * 100, 1) if calls > 0 else 0,
                'Sales': sales,
                'Conversion Rate %': round(sales / calls * 100, 2) if calls > 0 else 0,
                'Total Talk Time': sec_to_hms(r['total_talk_sec'] or 0),
                'Avg Talk Time': sec_to_hms(r['avg_talk_sec'] or 0),
                'Total Pause Time': sec_to_hms(r['total_pause_sec'] or 0),
                'Days Active': r['days_active'] or 0,
                'Calls per Day': round(calls / (r['days_active'] or 1), 1)
            })
        
        # Create filename
        exports_dir = ensure_exports_dir()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = exports_dir / f"agents_export_{timestamp}.csv"
        
        # Write CSV
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            if formatted_results:
                writer = csv.DictWriter(f, fieldnames=formatted_results[0].keys())
                writer.writeheader()
                writer.writerows(formatted_results)
        
        print_success(f"✅ Exported {len(formatted_results)} agents to: {filename}")
        
        # Ask to open folder
        open_choice = input("\nOpen exports folder? (y/N): ").strip().lower()
        if open_choice == 'y':
            os.startfile(exports_dir)
            
    except Exception as e:
        print_error(f"Export failed: {e}")

def export_campaigns_to_csv():
    """Export campaign performance to CSV"""
    print_header("📊 EXPORT CAMPAIGN PERFORMANCE TO CSV", Colors.CYAN)
    
    # Get date range
    start_date, end_date, range_text = get_date_range()
    
    print(f"\n📊 Exporting campaign data for {range_text}...")
    
    query = """
    SELECT 
        campaign_id,
        COUNT(*) as total_calls,
        SUM(CASE WHEN length_in_sec >= 5 THEN 1 ELSE 0 END) as answered,
        SUM(CASE WHEN term_reason IN ('ABANDON', 'QUEUETIMEOUT', 'NOAGENT') 
                 OR (length_in_sec = 0 AND queue_seconds > 0) THEN 1 ELSE 0 END) as abandoned,
        AVG(queue_seconds) as avg_queue_sec,
        AVG(length_in_sec) as avg_talk_sec,
        SUM(length_in_sec) as total_talk_sec,
        COUNT(DISTINCT user) as unique_agents
    FROM vicidial_closer_log
    WHERE call_date BETWEEN %s AND %s
    GROUP BY campaign_id
    ORDER BY total_calls DESC
    """
    
    try:
        results = db.execute_query(query, (start_date, end_date))
        
        if not results:
            print_warning("No campaign data found for selected period")
            return
        
        # Format results
        formatted_results = []
        for r in results:
            calls = r['total_calls'] or 0
            answered = r['answered'] or 0
            abandoned = r['abandoned'] or 0
            
            formatted_results.append({
                'Campaign': r['campaign_id'],
                'Total Calls': calls,
                'Answered': answered,
                'Answer Rate %': round(answered / calls * 100, 1) if calls > 0 else 0,
                'Abandoned': abandoned,
                'Abandon Rate %': round(abandoned / calls * 100, 1) if calls > 0 else 0,
                'Avg Queue (s)': round(r['avg_queue_sec'] or 0, 1),
                'Avg Talk Time': sec_to_hms(r['avg_talk_sec'] or 0),
                'Total Talk Time': sec_to_hms(r['total_talk_sec'] or 0),
                'Unique Agents': r['unique_agents'] or 0
            })
        
        # Create filename
        exports_dir = ensure_exports_dir()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = exports_dir / f"campaigns_export_{timestamp}.csv"
        
        # Write CSV
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            if formatted_results:
                writer = csv.DictWriter(f, fieldnames=formatted_results[0].keys())
                writer.writeheader()
                writer.writerows(formatted_results)
        
        print_success(f"✅ Exported {len(formatted_results)} campaigns to: {filename}")
        
        # Ask to open folder
        open_choice = input("\nOpen exports folder? (y/N): ").strip().lower()
        if open_choice == 'y':
            os.startfile(exports_dir)
            
    except Exception as e:
        print_error(f"Export failed: {e}")

# =============================================================================
# Excel Export Functions
# =============================================================================

def export_agent_performance(filename=None, period_days=30):
    """Export agent performance to Excel"""
    if not EXCEL_AVAILABLE:
        print_error("Excel export requires openpyxl. Run: pip install openpyxl")
        return None
    
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"agent_performance_{timestamp}.xlsx"
    
    exports_dir = ensure_exports_dir()
    filepath = exports_dir / filename
    
    print(f"\n📊 Exporting agent performance for last {period_days} days...")
    
    try:
        # Get agent summary
        summary_query = """
        SELECT 
            a.user,
            u.full_name,
            COUNT(*) as total_calls,
            SUM(CASE WHEN a.talk_sec >= 5 THEN 1 ELSE 0 END) as answered,
            SUM(a.talk_sec) as total_talk,
            AVG(CASE WHEN a.talk_sec >= 5 THEN a.talk_sec END) as avg_talk,
            MIN(a.event_time) as first_call,
            MAX(a.event_time) as last_call
        FROM vicidial_agent_log a
        LEFT JOIN vicidial_users u ON a.user = u.user
        WHERE a.event_time >= DATE_SUB(NOW(), INTERVAL %s DAY)
        GROUP BY a.user
        ORDER BY total_calls DESC
        """
        
        summary = db.execute_query(summary_query, (period_days,))
        
        # Get daily breakdown
        daily_query = """
        SELECT 
            a.user,
            u.full_name,
            DATE(a.event_time) as call_date,
            COUNT(*) as daily_calls,
            SUM(a.talk_sec) as daily_talk
        FROM vicidial_agent_log a
        LEFT JOIN vicidial_users u ON a.user = u.user
        WHERE a.event_time >= DATE_SUB(NOW(), INTERVAL %s DAY)
        GROUP BY a.user, DATE(a.event_time)
        ORDER BY call_date DESC, a.user
        """
        
        daily = db.execute_query(daily_query, (period_days,))
        
        # Get campaign breakdown
        campaign_query = """
        SELECT 
            a.user,
            u.full_name,
            c.campaign_id,
            COUNT(*) as campaign_calls,
            SUM(a.talk_sec) as campaign_talk
        FROM vicidial_agent_log a
        LEFT JOIN vicidial_users u ON a.user = u.user
        JOIN vicidial_closer_log c ON a.uniqueid = c.uniqueid
        WHERE a.event_time >= DATE_SUB(NOW(), INTERVAL %s DAY)
        GROUP BY a.user, c.campaign_id
        ORDER BY a.user, campaign_calls DESC
        """
        
        campaigns = db.execute_query(campaign_query, (period_days,))
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Summary sheet
            if summary:
                df_summary = pd.DataFrame(summary)
                df_summary['answer_rate'] = (df_summary['answered'] / df_summary['total_calls'] * 100).round(1)
                df_summary['total_talk'] = df_summary['total_talk'].apply(sec_to_hms)
                df_summary['avg_talk'] = df_summary['avg_talk'].apply(sec_to_hms)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Daily sheet
            if daily:
                df_daily = pd.DataFrame(daily)
                df_daily['daily_talk'] = df_daily['daily_talk'].apply(sec_to_hms)
                df_daily.to_excel(writer, sheet_name='Daily Breakdown', index=False)
            
            # Campaign sheet
            if campaigns:
                df_campaigns = pd.DataFrame(campaigns)
                df_campaigns['campaign_talk'] = df_campaigns['campaign_talk'].apply(sec_to_hms)
                df_campaigns.to_excel(writer, sheet_name='Campaign Breakdown', index=False)
        
        # Apply formatting
        format_excel_file(filepath)
        
        print_success(f"✅ Exported to: {filepath}")
        return str(filepath)
        
    except Exception as e:
        print_error(f"Export failed: {e}")
        return None

def export_campaign_performance(filename=None, period_days=30):
    """Export campaign performance to Excel"""
    if not EXCEL_AVAILABLE:
        print_error("Excel export requires openpyxl")
        return None
    
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"campaign_performance_{timestamp}.xlsx"
    
    exports_dir = ensure_exports_dir()
    filepath = exports_dir / filename
    
    print(f"\n📊 Exporting campaign performance for last {period_days} days...")
    
    try:
        # Campaign summary
        query = """
        SELECT 
            campaign_id,
            COUNT(*) as total_calls,
            SUM(CASE WHEN length_in_sec >= 5 THEN 1 ELSE 0 END) as answered,
            SUM(CASE WHEN term_reason IN ('ABANDON', 'QUEUETIMEOUT', 'NOAGENT') 
                     OR (length_in_sec = 0 AND queue_seconds > 0) THEN 1 ELSE 0 END) as abandoned,
            AVG(queue_seconds) as avg_queue,
            AVG(CASE WHEN length_in_sec >= 5 THEN length_in_sec END) as avg_talk,
            SUM(length_in_sec) as total_talk,
            MIN(call_date) as first_call,
            MAX(call_date) as last_call
        FROM vicidial_closer_log
        WHERE call_date >= DATE_SUB(NOW(), INTERVAL %s DAY)
        GROUP BY campaign_id
        ORDER BY total_calls DESC
        """
        
        campaigns = db.execute_query(query, (period_days,))
        
        if campaigns:
            df = pd.DataFrame(campaigns)
            df['answer_rate'] = (df['answered'] / df['total_calls'] * 100).round(1)
            df['abandon_rate'] = (df['abandoned'] / df['total_calls'] * 100).round(1)
            df['total_talk'] = df['total_talk'].apply(sec_to_hms)
            df['avg_talk'] = df['avg_talk'].apply(sec_to_hms)
            
            df.to_excel(filepath, index=False)
            format_excel_file(filepath)
            
            print_success(f"✅ Exported to: {filepath}")
            return str(filepath)
        else:
            print_warning("No campaign data found")
            return None
            
    except Exception as e:
        print_error(f"Export failed: {e}")
        return None

def export_queue_analysis(filename=None, period_days=7):
    """Export queue analysis to Excel"""
    if not EXCEL_AVAILABLE:
        print_error("Excel export requires openpyxl")
        return None
    
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"queue_analysis_{timestamp}.xlsx"
    
    exports_dir = ensure_exports_dir()
    filepath = exports_dir / filename
    
    print(f"\n📊 Exporting queue analysis for last {period_days} days...")
    
    try:
        # Hourly queue stats
        hourly_query = """
        SELECT 
            HOUR(call_date) as hour,
            COUNT(*) as calls,
            AVG(queue_seconds) as avg_queue,
            MAX(queue_seconds) as max_queue,
            SUM(CASE WHEN queue_seconds > 30 THEN 1 ELSE 0 END) as sl_violations
        FROM vicidial_closer_log
        WHERE call_date >= DATE_SUB(NOW(), INTERVAL %s DAY)
        GROUP BY HOUR(call_date)
        ORDER BY hour
        """
        
        hourly = db.execute_query(hourly_query, (period_days,))
        
        # Daily queue stats
        daily_query = """
        SELECT 
            DATE(call_date) as call_date,
            COUNT(*) as calls,
            AVG(queue_seconds) as avg_queue,
            MAX(queue_seconds) as max_queue
        FROM vicidial_closer_log
        WHERE call_date >= DATE_SUB(NOW(), INTERVAL %s DAY)
        GROUP BY DATE(call_date)
        ORDER BY call_date DESC
        """
        
        daily = db.execute_query(daily_query, (period_days,))
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            if hourly:
                df_hourly = pd.DataFrame(hourly)
                df_hourly['sl_compliance'] = ((df_hourly['calls'] - df_hourly['sl_violations']) / df_hourly['calls'] * 100).round(1)
                df_hourly.to_excel(writer, sheet_name='Hourly Analysis', index=False)
            
            if daily:
                df_daily = pd.DataFrame(daily)
                df_daily.to_excel(writer, sheet_name='Daily Analysis', index=False)
        
        format_excel_file(filepath)
        print_success(f"✅ Exported to: {filepath}")
        return str(filepath)
        
    except Exception as e:
        print_error(f"Export failed: {e}")
        return None

# =============================================================================
# Excel Export Sub-Menu
# =============================================================================

def excel_export_menu():
    """Menu for Excel exports"""
    while True:
        print_header("📊 EXCEL EXPORT", Colors.CYAN)
        print("  1. 👥 Export Agent Performance")
        print("  2. 📋 Export Campaign Performance")
        print("  3. ⏱️ Export Queue Analysis")
        print("  0. 🔙 Back")
        print("-" * 60)
        
        choice = input(f"\n{Colors.CYAN}Choice: {Colors.RESET}").strip()
        
        if choice == '1':
            print("\nSelect period:")
            print("  1. Last 7 days")
            print("  2. Last 30 days")
            print("  3. Last 90 days")
            period = input("Choice (1-3): ").strip()
            
            days = {'1': 7, '2': 30, '3': 90}.get(period, 30)
            export_agent_performance(period_days=days)
            input("\nPress Enter to continue...")
        
        elif choice == '2':
            print("\nSelect period:")
            print("  1. Last 7 days")
            print("  2. Last 30 days")
            print("  3. Last 90 days")
            period = input("Choice (1-3): ").strip()
            
            days = {'1': 7, '2': 30, '3': 90}.get(period, 30)
            export_campaign_performance(period_days=days)
            input("\nPress Enter to continue...")
        
        elif choice == '3':
            print("\nSelect period:")
            print("  1. Last 7 days")
            print("  2. Last 30 days")
            period = input("Choice (1-2): ").strip()
            
            days = {'1': 7, '2': 30}.get(period, 7)
            export_queue_analysis(period_days=days)
            input("\nPress Enter to continue...")
        
        elif choice == '0':
            break
        
        else:
            print_error("Invalid choice")
            input("\nPress Enter to continue...")

# =============================================================================
# Main Export Menu (for main.py integration)
# =============================================================================

def export_menu():
    """Main export menu - Called from main.py option 1 in Reports & Exports"""
    while True:
        print_header("📤 EXPORT TO CSV/EXCEL", Colors.CYAN)
        print("  1. 📊 Export Calls to CSV")
        print("  2. 👥 Export Agent Performance to CSV")
        print("  3. 📋 Export Campaign Performance to CSV")
        print("  4. 📈 Export to Excel (Formatted)")
        print("  5. 📁 Open Exports Folder")
        print("  0. 🔙 Back")
        print("-" * 70)
        print("   CSV: Raw data, Excel: Formatted reports with multiple sheets")
        print("-" * 70)
        
        choice = input(f"\n{Colors.CYAN}Choice: {Colors.RESET}").strip()
        
        if choice == '1':
            export_calls_to_csv()
        elif choice == '2':
            export_agents_to_csv()
        elif choice == '3':
            export_campaigns_to_csv()
        elif choice == '4':
            excel_export_menu()
        elif choice == '5':
            exports_dir = ensure_exports_dir()
            print(f"\n📁 Opening: {exports_dir}")
            try:
                os.startfile(exports_dir)
                print_success("✅ Folder opened")
            except:
                print_error("Could not open folder")
            input("\nPress Enter to continue...")
        elif choice == '0':
            break
        else:
            print_error("Invalid choice")
            input("\nPress Enter to continue...")

if __name__ == "__main__":
    export_menu()